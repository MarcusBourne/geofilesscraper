import os
import re
import asyncio
import threading
import queue
from urllib.parse import urljoin
from datetime import datetime

import requests
from bs4 import BeautifulSoup
from playwright.async_api import (
    async_playwright,
    TimeoutError as PlaywrightTimeoutError,
    Error as PlaywrightError,
)

from openpyxl import load_workbook
import tkinter as tk
from tkinter import font as tkfont
from tkinter import ttk, scrolledtext, messagebox

BASE_URL       = "https://gis.gov.nl.ca/minesen/geofiles/"
ENTRY_PATH     = "default.asp"
EXTERNAL_PREFIX= "https://www.gov.nl.ca/iet/mines-geoscience-reports-maps-docs"
USER_AGENT     = "PDF-Scraper/1.0"
SKIP_KEYWORDS  = ("map", "research")
CREDS_FILE     = "creds.txt"
XLSX_FILE      = "geomaster.xlsx"
REQUEST_TIMEOUT= 500
RESUME_FILE    = "resume.txt"
MISSING_FILE   = "filenames.txt"

log_queue   = queue.Queue()
pause_event = threading.Event()

def gui_log(msg):
    log_queue.put(msg)

def record_missing(name: str):
    with open(MISSING_FILE, "a") as f:
        f.write(f"{name}\n")

def load_allowed_ids(path=XLSX_FILE):
    if not os.path.exists(path):
        gui_log(f"Excel file not found: {path}")
        return set()
    wb = load_workbook(path, read_only=True, data_only=True)
    sheet = wb.active
    header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    idxs = { name: header.index(name) for name in (
        'MasterNo','Geofile_No','MasterNo2','Geofile_No2','MasterNo3','Geofile_No3'
    ) if name in header }
    raw = set()
    for row in sheet.iter_rows(min_row=2, values_only=True):
        for idx in idxs.values():
            if idx < len(row) and row[idx]:
                raw.add(str(row[idx]).strip())
    allowed = set()
    for rid in raw:
        allowed.add(rid)
        if '/' in rid:
            allowed.update({ rid.replace('/', '_'), rid.replace('/', '') })
            parts = rid.split('/')
            if len(parts) == 3:
                allowed.update({ f"{parts[0]}_{parts[2]}", f"{parts[0]}{parts[2]}" })
    return allowed

def get_resume_page():
    if os.path.exists(RESUME_FILE):
        try:
            return int(open(RESUME_FILE).read().split(',')[0].strip())
        except Exception:
            return 1
    return 1

def is_allowed(url: str) -> bool:
    name = os.path.basename(url)
    return any(p.search(name) or p.search(url) for p in ALLOWED_PATTERNS)

def upload_to_s3(url: str):
    filename = os.path.basename(url)
    bold_name = f"**{filename}**"

    if any(kw in filename.lower() for kw in SKIP_KEYWORDS) or not is_allowed(url):
        gui_log(f"Filtered out: {bold_name} not in database.")
        return

    downloads_dir = "downloads"
    os.makedirs(downloads_dir, exist_ok=True)
    filepath = os.path.join(downloads_dir, filename)

    if os.path.exists(filepath):
        gui_log(f"Skipping {bold_name}, already downloaded locally.")
        return

    gui_log(f"Downloading: {bold_name}")
    r = requests.get(url, stream=True,
                     headers={'User-Agent': USER_AGENT},
                     timeout=REQUEST_TIMEOUT)
    r.raise_for_status()
    with open(filepath, 'wb') as f:
        for chunk in r.iter_content(chunk_size=8192):
            if chunk:
                f.write(chunk)

    gui_log(f"Saved: {bold_name} to {downloads_dir}/")

def process(html: str):
    soup = BeautifulSoup(html, 'lxml')
    for a in soup.find_all('a', href=True):
        href = a['href']
        full = href if href.startswith('http') else urljoin(BASE_URL, href)
        if full.lower().endswith(('.pdf', '.zip')):
            upload_to_s3(full)
        elif full.startswith(EXTERNAL_PREFIX) and is_allowed(full):
            got_files = False
            try:
                resp = requests.get(full,
                                     headers={'User-Agent': USER_AGENT},
                                     timeout=REQUEST_TIMEOUT)
                resp.raise_for_status()
                ext_soup = BeautifulSoup(resp.text, 'lxml')
                for b in ext_soup.find_all('a', href=True):
                    link = b['href']
                    u = link if link.startswith('http') else urljoin(full, link)
                    if u.lower().endswith(('.pdf', '.zip')):
                        upload_to_s3(u)
                        got_files = True
                if not got_files:
                    mn = os.path.basename(full.rstrip('/'))
                    gui_log(f"No download for **{mn}**, logging.")
                    record_missing(mn)
            except Exception as e:
                gui_log(f"Error fetching external {full}: {e}")


ALLOWED_IDS = load_allowed_ids()
ALLOWED_PATTERNS = [
    re.compile(rf'(?<![A-Za-z0-9]){re.escape(i)}(?![A-Za-z0-9])',
               re.IGNORECASE)
    for i in ALLOWED_IDS
]

async def scraper_main():
    open(MISSING_FILE, 'w').close()
    entry = urljoin(BASE_URL, ENTRY_PATH)
    gui_log("\n========================= ✓ Starting Scraper ✓ =========================\n")
    gui_log(f"🔗 Entering Webpage: {BASE_URL} 🔗\n")
    async with async_playwright() as pw:
        browser = await pw.chromium.launch(headless=False)
        page = await browser.new_page(user_agent=USER_AGENT)

        await page.goto(entry)
        await page.wait_for_selector("form[name=searchForm]")
        await page.wait_for_timeout(2000)
        await page.evaluate("document.forms['searchForm'].submit()")
        await page.wait_for_load_state('networkidle')
        await page.wait_for_timeout(2000)

        resume_page = get_resume_page()
        if resume_page > 1:
            gui_log(f"🔄 Resuming from page {resume_page} 🔄")
            try:
                await page.evaluate(f"goPage({resume_page}, 'display.asp')")
            except PlaywrightError:
                gui_log(f"⚠️ goPage not available for resuming at page {resume_page}")
            await page.wait_for_load_state('networkidle')
            await page.wait_for_timeout(2000)

        soup0 = BeautifulSoup(await page.content(), 'lxml')
        last = soup0.select_one("img[src*='last.gif']")
        total = int(
            re.search(r"goPage\(\s*(\d+)", last.parent['href']).group(1)
            if last and last.parent and last.parent.has_attr('href') else 1
        )
        gui_log(f"📄 Total pages found: {total} 📄")

        for i in range(resume_page, total + 1):
            gui_log(f"\nScraping Page: {i}/{total} \n")

            html = ""
            for attempt in range(2):
                try:
                    await page.wait_for_load_state('networkidle', timeout=15000)
                    html = await page.content()
                    break
                except (PlaywrightTimeoutError, PlaywrightError) as e:
                    if attempt == 0:
                        gui_log(f"⚠️ Issue fetching content on page {i}, retrying in 2 seconds…")
                        await asyncio.sleep(2)
                    else:
                        gui_log(f"❌ Still failing to retrieve content for page {i}: {e}")
            process(html)

            now = datetime.now()
            with open(RESUME_FILE, 'w') as f:
                f.write(f"{i}, time: {now:%I:%M %p}, date: {now:%Y-%m-%d}")

            if i < total:
                next_page = i + 1
                try:
                    await page.evaluate(f"goPage({next_page}, 'display.asp')")
                    await page.wait_for_load_state('networkidle')
                    await page.wait_for_timeout(2000)
                    continue
                except PlaywrightError as e:
                    gui_log(f"⚠️ goPage eval failed for page {next_page}: {e}")

                try:
                    anchor = await page.query_selector("a:has(img[src*='next.gif'])")
                    if anchor:
                        href = await anchor.get_attribute('href')
                        nav_url = urljoin(BASE_URL, href)
                        await page.goto(nav_url)
                        await page.wait_for_load_state('networkidle')
                        await page.wait_for_timeout(2000)
                        continue
                except Exception as e:
                    gui_log(f"⚠️ Anchor href navigation failed for page {next_page}: {e}")

                try:
                    anchor = await page.query_selector("a:has(img[src*='next.gif'])")
                    if anchor:
                        await anchor.click()
                        await page.wait_for_load_state('networkidle')
                        await page.wait_for_timeout(2000)
                        continue
                except Exception as e:
                    gui_log(f"⚠️ Anchor click fallback failed for page {next_page}: {e}")

                try:
                    gui_log(f"🔄 Reloading and retrying goPage for {next_page}")
                    await page.reload()
                    await page.wait_for_timeout(2000)
                    await page.evaluate(f"goPage({next_page}, 'display.asp')")
                    await page.wait_for_load_state('networkidle')
                    await page.wait_for_timeout(2000)
                    continue
                except Exception as e:
                    gui_log(f"❌ Reload fallback failed for page {next_page}: {e}")

                try:
                    gui_log(f"♻️ Restarting browser for page {next_page}")
                    await browser.close()
                    browser = await pw.chromium.launch(headless=False)
                    page = await browser.new_page(user_agent=USER_AGENT)
                    await page.goto(entry)
                    await page.wait_for_selector("form[name=searchForm]")
                    await page.evaluate("document.forms['searchForm'].submit()")
                    await page.wait_for_load_state('networkidle')
                    await page.wait_for_timeout(2000)
                    await page.evaluate(f"goPage({next_page}, 'display.asp')")
                    await page.wait_for_load_state('networkidle')
                    await page.wait_for_timeout(2000)
                    continue
                except Exception as e:
                    gui_log(f"❌ Browser restart failed for page {next_page}: {e}")

                gui_log(f"❌ Cannot navigate to page {next_page}, stopping.")
                break

        await browser.close()
    gui_log("======================= ✓ Scraping Complete ✓ =======================")

def run_scraper():
    asyncio.run(scraper_main())

def start_gui():
    root = tk.Tk()
    root.title("Mines and Energy Geofiles Scraper")
    root.geometry("800x600")

    logo_img = tk.PhotoImage(file="cna.png").subsample(2, 2)
    frame = ttk.Frame(root, padding=10)
    frame.pack(fill=tk.BOTH, expand=True)
    logo_label = tk.Label(frame, image=logo_img)
    logo_label.image = logo_img
    logo_label.pack(pady=(0, 10))
    tk.Label(frame, text="Geofiles Scraper", font=("Georgia", 30, "bold")).pack()

    btn_frame = ttk.Frame(frame)
    btn_frame.pack(pady=10)

    tk.Button(
        btn_frame,
        text="Start Scraper",
        font=("Georgia", 10),
        command=lambda: threading.Thread(target=run_scraper, daemon=True).start()
    ).pack(side=tk.LEFT, padx=5, ipadx=20, ipady=10)

    tk.Button(
        btn_frame,
        text="Exit Scraper",
        font=("Georgia", 10),
        command=root.destroy
    ).pack(side=tk.LEFT, padx=5, ipadx=20, ipady=10)

    log_widget = scrolledtext.ScrolledText(frame, state='disabled', wrap=tk.WORD)
    bold_font = tkfont.nametofont(log_widget.cget("font")).copy()
    bold_font.configure(weight="bold")
    log_widget.tag_configure('bold', font=bold_font)
    log_widget.pack(fill=tk.BOTH, expand=True)

    def update_log():
        while not log_queue.empty():
            msg = log_queue.get()
            parts = msg.split('**')
            log_widget.config(state='normal')
            for idx, part in enumerate(parts):
                tag = 'bold' if idx % 2 else None
                log_widget.insert(tk.END, part, tag)
            log_widget.insert(tk.END, '\n')
            log_widget.see(tk.END)
            log_widget.config(state='disabled')
        root.after(100, update_log)

    root.after(100, update_log)
    root.mainloop()

if __name__ == '__main__':
    start_gui()